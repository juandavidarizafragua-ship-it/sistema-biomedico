import streamlit as st
import requests
from datetime import datetime, date, timedelta
import hashlib
import pandas as pd
import io

# -------- CONFIG --------
SUPABASE_URL = "https://hivvykyslqodrrfmteer.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImhpdnZ5a3lzbHFvZHJyZm10ZWVyIiwicm9sZSI6ImFub24iLCJpYXQiOjE3Nzg3MTA0NzQsImV4cCI6MjA5NDI4NjQ3NH0.1dBUFD9myAK9057E0g6RVFKellm6RT_6E15RHz63ryc"
HEADERS = {
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "apikey": SUPABASE_KEY,
    "Content-Type": "application/json"
}

NOMBRE_EMPRESA = "Biomédica Demo"
COLOR = "#2E86AB"
SLUG = "biodemo"

USUARIOS = {
    "admin": {
        "password": hashlib.sha256("admin123".encode()).hexdigest(),
        "rol": "admin"
    }
}

T_CLIENTES = f"{SLUG}_clientes"
T_ALERTAS = f"{SLUG}_alertas"
T_EQUIPOS = f"{SLUG}_equipos"
T_HOJAS = f"{SLUG}_hojas_vida"
T_CRONOGRAMA = f"{SLUG}_cronograma"

# -------- SUPABASE HELPERS --------
def sb_get(tabla, params=""):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/{tabla}?{params}", headers=HEADERS)
    return r.json() if r.status_code == 200 else []

def sb_post(tabla, datos):
    r = requests.post(f"{SUPABASE_URL}/rest/v1/{tabla}", headers={**HEADERS, "Prefer": "return=representation"}, json=datos)
    return r.status_code == 201

def sb_patch(tabla, id_val, datos):
    r = requests.patch(f"{SUPABASE_URL}/rest/v1/{tabla}?id=eq.{id_val}", headers={**HEADERS, "Prefer": "return=minimal"}, json=datos)
    return r.status_code == 204

def subir_documento(archivo, carpeta=None):
    if carpeta is None:
        carpeta = SLUG
    try:
        url = f"{SUPABASE_URL}/storage/v1/object/documentos/{carpeta}/{archivo.name}"
        h = {"Authorization": f"Bearer {SUPABASE_KEY}", "apikey": SUPABASE_KEY, "Content-Type": archivo.type}
        r = requests.post(url, headers=h, data=archivo.getvalue())
        if r.status_code in [200, 201]:
            return f"{SUPABASE_URL}/storage/v1/object/public/documentos/{carpeta}/{archivo.name}"
        return None
    except:
        return None

def listar_documentos(carpeta=None):
    if carpeta is None:
        carpeta = SLUG
    try:
        url = f"{SUPABASE_URL}/storage/v1/object/list/documentos"
        r = requests.post(url, headers={**HEADERS}, json={"prefix": carpeta + "/", "limit": 100})
        return r.json() if r.status_code == 200 else []
    except:
        return []

# -------- GENERAR HOJA DE VIDA EXCEL --------
def generar_hoja_vida_excel(equipo, historial):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import openpyxl

    wb = openpyxl.Workbook()

    # === HOJA 1: DATOS DEL EQUIPO ===
    ws1 = wb.active
    ws1.title = "Datos del Equipo"

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
    title_font = Font(name="Arial", bold=True, size=14, color="2E86AB")
    label_font = Font(name="Arial", bold=True, size=10, color="333333")
    value_font = Font(name="Arial", size=10, color="555555")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )
    section_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 35
    ws1.column_dimensions["C"].width = 28
    ws1.column_dimensions["D"].width = 35

    # Título
    ws1.merge_cells("A1:D1")
    ws1["A1"] = "HOJA DE VIDA DE EQUIPO BIOMÉDICO"
    ws1["A1"].font = title_font
    ws1["A1"].alignment = Alignment(horizontal="center")

    ws1.merge_cells("A2:D2")
    ws1["A2"] = f"{NOMBRE_EMPRESA} — Generado {date.today().strftime('%d/%m/%Y')}"
    ws1["A2"].font = Font(name="Arial", size=10, color="888888")
    ws1["A2"].alignment = Alignment(horizontal="center")

    # Sección: Registro Histórico
    row = 4
    ws1.merge_cells(f"A{row}:D{row}")
    ws1[f"A{row}"] = "REGISTRO HISTÓRICO"
    ws1[f"A{row}"].font = header_font
    ws1[f"A{row}"].fill = header_fill

    datos_equipo = [
        ("Nombre del equipo", equipo.get("nombre", ""), "Marca", equipo.get("marca", "")),
        ("Modelo", equipo.get("modelo", ""), "Serie", equipo.get("serie", "")),
        ("Código inventario", equipo.get("id", ""), "Ubicación", equipo.get("ubicacion", "")),
        ("Fecha de adquisición", equipo.get("fecha_adquisicion", ""), "Estado", equipo.get("estado", "")),
        ("Registro INVIMA", equipo.get("registro_invima", ""), "Clase de riesgo", equipo.get("clasificacion_riesgo", "")),
        ("Frecuencia mantenimiento", equipo.get("frecuencia_mantenimiento", ""), "Cliente", equipo.get("cliente_nombre", "")),
    ]

    for i, (l1, v1, l2, v2) in enumerate(datos_equipo):
        r = row + 1 + i
        ws1[f"A{r}"] = l1
        ws1[f"A{r}"].font = label_font
        ws1[f"A{r}"].border = border
        ws1[f"B{r}"] = str(v1)
        ws1[f"B{r}"].font = value_font
        ws1[f"B{r}"].border = border
        ws1[f"C{r}"] = l2
        ws1[f"C{r}"].font = label_font
        ws1[f"C{r}"].border = border
        ws1[f"D{r}"] = str(v2)
        ws1[f"D{r}"].font = value_font
        ws1[f"D{r}"].border = border

    # Sección: Información adicional
    row = row + len(datos_equipo) + 2
    ws1.merge_cells(f"A{row}:D{row}")
    ws1[f"A{row}"] = "INFORMACIÓN ADICIONAL"
    ws1[f"A{row}"].font = header_font
    ws1[f"A{row}"].fill = header_fill

    info_adicional = [
        ("Notas / Observaciones", equipo.get("notas", "—")),
        ("Fecha de registro en sistema", equipo.get("created_at", "")[:10] if equipo.get("created_at") else ""),
        ("Total mantenimientos registrados", str(len(historial))),
    ]

    for i, (label, val) in enumerate(info_adicional):
        r = row + 1 + i
        ws1[f"A{r}"] = label
        ws1[f"A{r}"].font = label_font
        ws1[f"A{r}"].border = border
        ws1.merge_cells(f"B{r}:D{r}")
        ws1[f"B{r}"] = str(val)
        ws1[f"B{r}"].font = value_font
        ws1[f"B{r}"].border = border

    # === HOJA 2: HISTORIAL DE MANTENIMIENTO ===
    ws2 = wb.create_sheet("Historial de Mantenimiento")

    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"REGISTRO HISTÓRICO DE MANTENIMIENTO — {equipo.get('nombre', '')} ({equipo.get('marca', '')} {equipo.get('modelo', '')})"
    ws2["A1"].font = title_font
    ws2["A1"].alignment = Alignment(horizontal="center")

    headers = ["Fecha", "Tipo de Mtto", "Trabajo Realizado", "Repuestos", "Observaciones", "Responsable"]
    anchos = [14, 16, 45, 30, 30, 20]

    for col, (header, ancho) in enumerate(zip(headers, anchos), 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
        ws2.column_dimensions[chr(64 + col)].width = ancho

    if historial:
        for i, h in enumerate(historial):
            r = 4 + i
            datos_fila = [
                h.get("fecha", ""),
                h.get("tipo_mantenimiento", ""),
                h.get("trabajo_realizado", ""),
                h.get("repuestos", ""),
                h.get("observaciones", ""),
                h.get("tecnico", "")
            ]
            for col, val in enumerate(datos_fila, 1):
                cell = ws2.cell(row=r, column=col, value=str(val or "—"))
                cell.font = value_font
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    else:
        ws2.merge_cells("A4:F4")
        ws2["A4"] = "No hay registros de mantenimiento"
        ws2["A4"].font = Font(name="Arial", size=10, color="999999")
        ws2["A4"].alignment = Alignment(horizontal="center")

    # Pie
    r_final = max(4 + len(historial), 5) + 2
    ws2.merge_cells(f"A{r_final}:F{r_final}")
    ws2[f"A{r_final}"] = f"NR: NO REGISTRA | NT: NO TIENE | NA: NO APLICA — Generado por {NOMBRE_EMPRESA} / A.R.I.Z.A."
    ws2[f"A{r_final}"].font = Font(name="Arial", size=8, color="AAAAAA")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# -------- GENERAR REPORTE MANTENIMIENTO PDF --------
def generar_reporte_pdf(equipo, mantenimiento):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=20*mm, rightMargin=20*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    story = []

    # Estilos personalizados
    style_title = ParagraphStyle("Title2", parent=styles["Normal"], fontSize=14, fontName="Helvetica-Bold", alignment=1, spaceAfter=6)
    style_subtitle = ParagraphStyle("Sub2", parent=styles["Normal"], fontSize=9, fontName="Helvetica", alignment=1, textColor=colors.gray, spaceAfter=12)
    style_section = ParagraphStyle("Sec2", parent=styles["Normal"], fontSize=10, fontName="Helvetica-Bold", spaceAfter=4, spaceBefore=8)
    style_normal = ParagraphStyle("Norm2", parent=styles["Normal"], fontSize=9, fontName="Helvetica", spaceAfter=2)
    style_small = ParagraphStyle("Small2", parent=styles["Normal"], fontSize=8, fontName="Helvetica", textColor=colors.gray)

    color_header = colors.HexColor("#1a1a1a")
    color_light = colors.HexColor("#f5f5f5")
    color_accent = colors.HexColor("#2E86AB")
    color_border = colors.HexColor("#CCCCCC")

    # ---- ENCABEZADO ----
    story.append(Paragraph("REPORTE DE MANTENIMIENTO", style_title))
    story.append(Paragraph(f"{NOMBRE_EMPRESA} — Powered by A.R.I.Z.A.", style_subtitle))

    # ---- INFO GENERAL ----
    fecha_reporte = mantenimiento.get("fecha_ejecucion", str(date.today()))
    reporte_num = f"RPT-{equipo.get('id', '0')}-{mantenimiento.get('id', '0')}"

    info_data = [
        ["REPORTE N°", reporte_num, "FECHA", fecha_reporte],
    ]
    info_table = Table(info_data, colWidths=[80, 150, 50, 150])
    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), color_header), ("TEXTCOLOR", (0, 0), (0, 0), colors.white),
        ("BACKGROUND", (2, 0), (2, 0), color_header), ("TEXTCOLOR", (2, 0), (2, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"), ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (0, 0), "Helvetica-Bold"), ("FONTNAME", (2, 0), (2, 0), "Helvetica-Bold"),
        ("BOX", (0, 0), (-1, -1), 0.5, color_border),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, color_border),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 10))

    # ---- INFORMACIÓN DEL EQUIPO ----
    story.append(Paragraph("INFORMACI\u00d3N DEL EQUIPO", style_section))

    eq_data = [
        ["NOMBRE", equipo.get("nombre", ""), "C\u00d3D. INVENTARIO", str(equipo.get("id", ""))],
        ["MARCA", equipo.get("marca", ""), "UBICACI\u00d3N", equipo.get("ubicacion", "")],
        ["MODELO", equipo.get("modelo", ""), "CLIENTE", equipo.get("cliente_nombre", "")],
        ["SERIE", equipo.get("serie", ""), "CLASE DE RIESGO", equipo.get("clasificacion_riesgo", "")],
        ["REG. INVIMA", equipo.get("registro_invima", ""), "FREC. MANTENIMIENTO", equipo.get("frecuencia_mantenimiento", "")],
    ]
    eq_table = Table(eq_data, colWidths=[80, 150, 100, 130])
    eq_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), color_light), ("BACKGROUND", (2, 0), (2, -1), color_light),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"), ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"), ("FONTNAME", (3, 0), (3, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOX", (0, 0), (-1, -1), 0.5, color_border),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, color_border),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(eq_table)
    story.append(Spacer(1, 10))

    # ---- TIPO DE MANTENIMIENTO ----
    story.append(Paragraph("TIPO DE MANTENIMIENTO", style_section))
    tipo = mantenimiento.get("tipo", "Preventivo")
    tipos = ["PREVENTIVO", "CORRECTIVO", "CALIBRACI\u00d3N", "INSPECCI\u00d3N"]
    tipo_row = []
    for t in tipos:
        marca = "[X]" if t == tipo.upper() else "[  ]"
        tipo_row.append(f"{marca} {t}")

    tipo_table = Table([tipo_row], colWidths=[115, 115, 115, 115])
    tipo_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"), ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOX", (0, 0), (-1, -1), 0.5, color_border),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, color_border),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(tipo_table)
    story.append(Spacer(1, 10))

    # ---- ACTIVIDADES ----
    story.append(Paragraph("ACTIVIDADES REALIZADAS", style_section))
    actividades_data = [
        ["MANTENIMIENTO / REPARACI\u00d3N", "VERIFICACI\u00d3N", "PRUEBAS"],
        ["[  ] Mtto. general", "[  ] Sensores / Transductores", "[  ] Prueba de funcionamiento"],
        ["[  ] Mtto. hidr\u00e1ulico/neum\u00e1tico", "[  ] Interfaces / Mangueras / Cables", "[  ] Prueba de fugas"],
        ["[  ] Mtto. el\u00e9ctrico/electr\u00f3nico", "[  ] Bombillos / Conectores", "[  ] Prueba de rendimiento"],
        ["[  ] Reparaci\u00f3n exterior", "[  ] Bater\u00edas / Cargadores / Fuentes", "[  ] Puesta a punto"],
        ["[  ] Mtto. mec\u00e1nico", "[  ] Controles / Interruptores / Fusibles", "[  ] Prueba de alarmas"],
        ["[  ] Reparaci\u00f3n el\u00e9ctrica", "[  ] M\u00f3dulo de impresi\u00f3n", ""],
        ["[  ] Reparaci\u00f3n hidr\u00e1ulica", "[  ] Software / Firmware", ""],
    ]
    act_table = Table(actividades_data, colWidths=[155, 165, 140])
    act_style = [
        ("BACKGROUND", (0, 0), (-1, 0), color_header), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"), ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("BOX", (0, 0), (-1, -1), 0.5, color_border),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, color_border),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    act_table.setStyle(TableStyle(act_style))
    story.append(act_table)
    story.append(Spacer(1, 10))

    # ---- DIAGNÓSTICO / DESCRIPCIÓN ----
    story.append(Paragraph("DIAGN\u00d3STICO / DESCRIPCI\u00d3N DE LA INTERVENCI\u00d3N", style_section))
    descripcion = mantenimiento.get("descripcion", "") or ""
    observaciones = mantenimiento.get("observaciones", "") or ""
    texto_diag = f"{descripcion}\n{observaciones}" if observaciones else descripcion

    diag_data = [[Paragraph(texto_diag or "Sin descripci\u00f3n registrada", style_normal)]]
    diag_table = Table(diag_data, colWidths=[460])
    diag_table.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, color_border),
        ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(diag_table)
    story.append(Spacer(1, 10))

    # ---- REPUESTOS ----
    story.append(Paragraph("REPUESTOS Y/O MATERIALES UTILIZADOS", style_section))
    repuestos_text = mantenimiento.get("repuestos", "") or "Ninguno"

    rep_data = [
        ["DESCRIPCI\u00d3N", "UNIDAD", "CANTIDAD", "REFERENCIA"],
        [repuestos_text, "", "", ""],
        ["", "", "", ""],
        ["", "", "", ""],
    ]
    rep_table = Table(rep_data, colWidths=[200, 80, 80, 100])
    rep_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), color_header), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"), ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("BOX", (0, 0), (-1, -1), 0.5, color_border),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, color_border),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(rep_table)
    story.append(Spacer(1, 10))

    # ---- CONDICIONES DE ENTREGA ----
    story.append(Paragraph("CONDICIONES DE ENTREGA", style_section))
    entrega_data = [["[  ] EN FUNCIONAMIENTO", "[  ] CON FALLA", "[  ] FUERA DE SERVICIO", "[  ] BAJA"]]
    ent_table = Table(entrega_data, colWidths=[130, 100, 130, 100])
    ent_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"), ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOX", (0, 0), (-1, -1), 0.5, color_border),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, color_border),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(ent_table)
    story.append(Spacer(1, 10))

    # ---- OBSERVACIONES ----
    story.append(Paragraph("OBSERVACIONES", style_section))
    obs_data = [[Paragraph(observaciones or "Sin observaciones adicionales", style_normal)]]
    obs_table = Table(obs_data, colWidths=[460])
    obs_table.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, color_border),
        ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 30),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(obs_table)
    story.append(Spacer(1, 15))

    # ---- FIRMAS ----
    story.append(Paragraph("INFORMACI\u00d3N DE ENTREGA", style_section))
    tecnico = mantenimiento.get("tecnico", "")
    firma_data = [
        ["NOMBRE INGENIERO/T\u00c9CNICO", "NOMBRE JEFE/COORDINADOR"],
        [tecnico, ""],
        ["", ""],
        ["FIRMA: ________________________", "FIRMA: ________________________"],
    ]
    firma_table = Table(firma_data, colWidths=[230, 230])
    firma_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"), ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("BOX", (0, 0), (-1, -1), 0.5, color_border),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, color_border),
        ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(firma_table)

    # Pie
    story.append(Spacer(1, 15))
    story.append(Paragraph(f"Generado autom\u00e1ticamente por {NOMBRE_EMPRESA} / A.R.I.Z.A. — {date.today().strftime('%d/%m/%Y')}", style_small))

    doc.build(story)
    return buffer.getvalue()


# -------- PAGE CONFIG --------
st.set_page_config(page_title=NOMBRE_EMPRESA, page_icon="🔬", layout="wide")

st.markdown(f"""
<style>
.stApp {{ background-color: #0a0a0a; }}
.stButton > button {{
    background: linear-gradient(135deg, #0a1a2a, #112233) !important;
    color: {COLOR} !important; border: 1px solid {COLOR}66 !important;
    border-radius: 4px !important; font-size: 12px !important;
}}
.stTextInput > div > div > input, .stNumberInput > div > div > input,
.stDateInput > div > div > input, .stSelectbox > div > div > div,
.stTextArea > div > div > textarea {{
    background-color: #111 !important; color: #e0e0e0 !important;
    border: 1px solid {COLOR}44 !important;
}}
.metric-card {{
    background: linear-gradient(135deg, #111, #1a1a1a);
    border: 1px solid {COLOR}33; border-radius: 8px;
    padding: 20px 24px; text-align: center;
}}
.metric-num {{ font-family: serif; font-size: 42px; font-weight: 700; color: {COLOR}; line-height: 1; }}
.metric-label {{ font-size: 11px; color: #888; letter-spacing: 2px; margin-top: 6px; }}
.divider {{ border:none; height:1px; background:linear-gradient(90deg,transparent,{COLOR}44,transparent); margin:20px 0; }}
</style>
""", unsafe_allow_html=True)

# -------- LOGIN --------
if "autenticado" not in st.session_state:
    st.session_state.autenticado = False

if not st.session_state.autenticado:
    st.markdown(f"""
    <div style='text-align:center; padding:60px 0 30px 0;'>
        <div style='font-family:serif; font-size:11px; color:{COLOR}; letter-spacing:4px;'>{NOMBRE_EMPRESA.upper()}</div>
        <div style='font-family:serif; font-size:28px; font-weight:700; color:#fff; margin:8px 0;'>Gesti\u00f3n de Equipos Biom\u00e9dicos</div>
        <div style='font-size:11px; color:#555; letter-spacing:2px;'>Powered by A.R.I.Z.A.</div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 1.5, 1])
    with col2:
        usr = st.text_input("Usuario", key="login_usr")
        pwd = st.text_input("Contrase\u00f1a", type="password", key="login_pwd")
        if st.button("Ingresar", use_container_width=True):
            pwd_hash = hashlib.sha256(pwd.encode()).hexdigest()
            if usr in USUARIOS and USUARIOS[usr]["password"] == pwd_hash:
                st.session_state.autenticado = True
                st.session_state.usuario = usr
                st.rerun()
            else:
                st.error("Credenciales incorrectas")
    st.stop()

# -------- HEADER --------
st.markdown(f"""
<div style='text-align:center; padding:20px 0 10px 0; border-bottom:1px solid {COLOR}33; margin-bottom:20px;'>
    <div style='font-family:serif; font-size:11px; color:{COLOR}; letter-spacing:4px;'>{NOMBRE_EMPRESA.upper()}</div>
    <div style='font-family:serif; font-size:24px; font-weight:700; color:#fff;'>Gesti\u00f3n de Equipos Biom\u00e9dicos</div>
</div>
""", unsafe_allow_html=True)

# -------- MENÚ --------
if "menu" not in st.session_state:
    st.session_state.menu = "Inventario"

modulos = ["Inventario", "Cronograma", "Hojas de Vida", "Clientes", "Alertas", "Documentos"]
cols = st.columns(len(modulos))
for i, mod in enumerate(modulos):
    with cols[i]:
        if st.button(f"\u25c8 {mod.upper()}", use_container_width=True, key=f"menu_{mod}"):
            st.session_state.menu = mod
            st.rerun()

menu = st.session_state.menu
st.markdown("<hr class='divider'>", unsafe_allow_html=True)

clientes_list = sb_get(T_CLIENTES, "order=nombre.asc")
nombres_clientes = ["Todos"] + [c.get("nombre", "") for c in clientes_list]


# ================================================================
# INVENTARIO
# ================================================================
if menu == "Inventario":
    equipos = sb_get(T_EQUIPOS, "order=created_at.desc")

    st.markdown(f"<div style='font-family:serif; font-size:20px; color:{COLOR};'>Inventario de Equipos</div>", unsafe_allow_html=True)

    activos = len([e for e in equipos if e.get("estado") == "Activo"])
    fuera = len([e for e in equipos if e.get("estado") == "Fuera de servicio"])

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f"<div class='metric-card'><div class='metric-num'>{len(equipos)}</div><div class='metric-label'>TOTAL</div></div>", unsafe_allow_html=True)
    with col2:
        st.markdown(f"<div class='metric-card'><div class='metric-num' style='color:#44ff88;'>{activos}</div><div class='metric-label'>ACTIVOS</div></div>", unsafe_allow_html=True)
    with col3:
        st.markdown(f"<div class='metric-card'><div class='metric-num' style='color:#ffaa44;'>{fuera}</div><div class='metric-label'>FUERA SERVICIO</div></div>", unsafe_allow_html=True)

    st.markdown("<hr class='divider'>", unsafe_allow_html=True)

    with st.expander("\u2795 Registrar equipo"):
        ec1, ec2 = st.columns(2)
        with ec1:
            eq_nombre = st.text_input("Nombre del equipo", key="eq_nombre")
            eq_marca = st.text_input("Marca", key="eq_marca")
            eq_modelo = st.text_input("Modelo", key="eq_modelo")
            eq_serie = st.text_input("N\u00famero de serie", key="eq_serie")
            eq_ubicacion = st.text_input("Ubicaci\u00f3n / \u00c1rea", key="eq_ubic")
        with ec2:
            eq_cliente = st.selectbox("Cliente", [c.get("nombre") for c in clientes_list] if clientes_list else ["Sin clientes"], key="eq_cliente")
            eq_clasificacion = st.selectbox("Clasificaci\u00f3n de riesgo", ["I", "IIA", "IIB", "III"], key="eq_clasif")
            eq_registro = st.text_input("Registro INVIMA", key="eq_invima")
            eq_fecha_adq = st.date_input("Fecha adquisici\u00f3n", key="eq_fecha")
            eq_frec = st.selectbox("Frecuencia mantenimiento", ["Mensual", "Bimestral", "Trimestral", "Semestral", "Anual"], key="eq_frec")

        eq_notas = st.text_area("Observaciones", key="eq_notas")

        if st.button("\u2713 Registrar equipo", key="btn_eq"):
            if eq_nombre and eq_marca:
                sb_post(T_EQUIPOS, {
                    "nombre": eq_nombre, "marca": eq_marca, "modelo": eq_modelo,
                    "serie": eq_serie, "ubicacion": eq_ubicacion,
                    "cliente_nombre": eq_cliente, "clasificacion_riesgo": eq_clasificacion,
                    "registro_invima": eq_registro, "fecha_adquisicion": str(eq_fecha_adq),
                    "estado": "Activo", "frecuencia_mantenimiento": eq_frec, "notas": eq_notas
                })
                st.success("\u2713 Equipo registrado")
                st.rerun()

    st.markdown("<hr class='divider'>", unsafe_allow_html=True)

    fc1, fc2 = st.columns(2)
    with fc1:
        filtro_cliente = st.selectbox("Filtrar por cliente:", nombres_clientes, key="inv_fc")
    with fc2:
        filtro_estado = st.selectbox("Filtrar por estado:", ["Todos", "Activo", "Fuera de servicio", "Dado de baja"], key="inv_fe")

    for equipo in equipos:
        if filtro_cliente != "Todos" and equipo.get("cliente_nombre") != filtro_cliente:
            continue
        if filtro_estado != "Todos" and equipo.get("estado") != filtro_estado:
            continue

        estado = equipo.get("estado", "Activo")
        color_est = {"Activo": "#44ff88", "Fuera de servicio": "#ffaa44", "Dado de baja": "#ff4444"}.get(estado, "#888")

        with st.expander(f"\U0001f52c {equipo.get('nombre')} \u2014 {equipo.get('marca')} {equipo.get('modelo')} \u2014 {equipo.get('cliente_nombre', '')}"):
            ic1, ic2 = st.columns(2)
            with ic1:
                st.markdown(f"**Nombre:** {equipo.get('nombre')}")
                st.markdown(f"**Marca:** {equipo.get('marca')}")
                st.markdown(f"**Modelo:** {equipo.get('modelo')}")
                st.markdown(f"**Serie:** {equipo.get('serie', '\u2014')}")
            with ic2:
                st.markdown(f"**Cliente:** {equipo.get('cliente_nombre', '\u2014')}")
                st.markdown(f"**Riesgo:** {equipo.get('clasificacion_riesgo', '\u2014')}")
                st.markdown(f"**INVIMA:** {equipo.get('registro_invima', '\u2014')}")
                st.markdown(f"**Frecuencia:** {equipo.get('frecuencia_mantenimiento', '\u2014')}")
                st.markdown(f"<span style='color:{color_est}'>**Estado: {estado}**</span>", unsafe_allow_html=True)

            ba, bb, bc = st.columns(3)
            with ba:
                if estado != "Activo" and st.button("Activar", key=f"eq_act_{equipo.get('id')}"):
                    sb_patch(T_EQUIPOS, equipo.get("id"), {"estado": "Activo"})
                    st.rerun()
            with bb:
                if estado == "Activo" and st.button("Fuera de servicio", key=f"eq_fs_{equipo.get('id')}"):
                    sb_patch(T_EQUIPOS, equipo.get("id"), {"estado": "Fuera de servicio"})
                    st.rerun()
            with bc:
                if estado != "Dado de baja" and st.button("Dar de baja", key=f"eq_baja_{equipo.get('id')}"):
                    sb_patch(T_EQUIPOS, equipo.get("id"), {"estado": "Dado de baja"})
                    st.rerun()

    if equipos and st.button("\U0001f4ca Exportar inventario", key="exp_inv"):
        df = pd.DataFrame(equipos)
        buf = io.BytesIO()
        df.to_excel(buf, index=False)
        st.download_button("Descargar", buf.getvalue(), f"inventario_{date.today()}.xlsx")


# ================================================================
# CRONOGRAMA
# ================================================================
elif menu == "Cronograma":
    cronograma = sb_get(T_CRONOGRAMA, "order=fecha_programada.asc")
    equipos = sb_get(T_EQUIPOS, "order=nombre.asc")

    st.markdown(f"<div style='font-family:serif; font-size:20px; color:{COLOR};'>Cronograma de Mantenimiento</div>", unsafe_allow_html=True)

    hoy = date.today()
    vencidos = [c for c in cronograma if c.get("fecha_programada") and c.get("fecha_programada") < str(hoy) and c.get("estado") != "Completado"]
    proximos = [c for c in cronograma if c.get("fecha_programada") and str(hoy) <= c.get("fecha_programada") <= str(hoy + timedelta(days=7)) and c.get("estado") != "Completado"]

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f"<div class='metric-card'><div class='metric-num' style='color:#ff4444;'>{len(vencidos)}</div><div class='metric-label'>VENCIDOS</div></div>", unsafe_allow_html=True)
    with col2:
        st.markdown(f"<div class='metric-card'><div class='metric-num' style='color:#ffaa44;'>{len(proximos)}</div><div class='metric-label'>PR\u00d3X. 7 D\u00cdAS</div></div>", unsafe_allow_html=True)
    with col3:
        st.markdown(f"<div class='metric-card'><div class='metric-num'>{len(cronograma)}</div><div class='metric-label'>TOTAL</div></div>", unsafe_allow_html=True)

    st.markdown("<hr class='divider'>", unsafe_allow_html=True)

    with st.expander("\u2795 Programar mantenimiento"):
        if equipos:
            opciones_eq = {f"{e.get('nombre')} \u2014 {e.get('marca')} {e.get('modelo')} ({e.get('cliente_nombre', '')})": e.get("id") for e in equipos}
            cr_equipo_sel = st.selectbox("Equipo", list(opciones_eq.keys()), key="cr_eq")
            cr_equipo_id = opciones_eq[cr_equipo_sel]
        else:
            st.warning("Registra equipos primero")
            cr_equipo_id = None

        cc1, cc2 = st.columns(2)
        with cc1:
            cr_tipo = st.selectbox("Tipo", ["Preventivo", "Correctivo", "Calibraci\u00f3n", "Inspecci\u00f3n"], key="cr_tipo")
            cr_fecha = st.date_input("Fecha programada", key="cr_fecha")
        with cc2:
            cr_tecnico = st.text_input("T\u00e9cnico responsable", key="cr_tec")
            cr_prioridad = st.selectbox("Prioridad", ["Normal", "Alta", "Urgente"], key="cr_pri")

        cr_desc = st.text_area("Descripci\u00f3n del trabajo", key="cr_desc")

        if st.button("\u2713 Programar", key="btn_cr"):
            if cr_equipo_id:
                nombre_eq = cr_equipo_sel.split(" \u2014 ")[0]
                sb_post(T_CRONOGRAMA, {
                    "equipo_id": cr_equipo_id, "equipo_nombre": nombre_eq,
                    "tipo": cr_tipo, "fecha_programada": str(cr_fecha),
                    "tecnico": cr_tecnico, "prioridad": cr_prioridad,
                    "descripcion": cr_desc, "estado": "Pendiente"
                })
                st.success("\u2713 Programado")
                st.rerun()

    st.markdown("<hr class='divider'>", unsafe_allow_html=True)

    filtro_cr = st.selectbox("Filtrar:", ["Pendientes", "Todos", "Completados"], key="cr_filtro")

    for item in cronograma:
        estado = item.get("estado", "Pendiente")
        fecha = item.get("fecha_programada", "")
        es_vencido = fecha and fecha < str(hoy) and estado != "Completado"

        if filtro_cr == "Pendientes" and estado != "Pendiente":
            continue
        if filtro_cr == "Completados" and estado != "Completado":
            continue

        icono = "\U0001f534" if es_vencido else ("\U0001f7e1" if estado == "Pendiente" else "\U0001f7e2")

        with st.expander(f"{icono} {item.get('equipo_nombre', '\u2014')} \u2014 {item.get('tipo')} \u2014 {fecha}"):
            st.markdown(f"**Equipo:** {item.get('equipo_nombre', '\u2014')}")
            st.markdown(f"**Tipo:** {item.get('tipo')}")
            st.markdown(f"**T\u00e9cnico:** {item.get('tecnico', '\u2014')}")
            st.markdown(f"**Descripci\u00f3n:** {item.get('descripcion', '\u2014')}")
            st.markdown(f"**Estado:** {estado}")

            if estado != "Completado":
                st.markdown("<hr class='divider'>", unsafe_allow_html=True)
                comp_obs = st.text_area("Observaciones del trabajo", key=f"cr_obs_{item.get('id')}")
                comp_rep = st.text_input("Repuestos utilizados", key=f"cr_rep_{item.get('id')}")
                comp_fecha = st.date_input("Fecha ejecuci\u00f3n", value=date.today(), key=f"cr_fr_{item.get('id')}")

                if st.button("\u2713 Completar mantenimiento", key=f"cr_comp_{item.get('id')}"):
                    sb_patch(T_CRONOGRAMA, item.get("id"), {
                        "estado": "Completado", "observaciones": comp_obs,
                        "repuestos": comp_rep, "fecha_ejecucion": str(comp_fecha)
                    })
                    sb_post(T_HOJAS, {
                        "equipo_id": item.get("equipo_id"), "equipo_nombre": item.get("equipo_nombre"),
                        "tipo_mantenimiento": item.get("tipo"), "fecha": str(comp_fecha),
                        "tecnico": item.get("tecnico"),
                        "trabajo_realizado": (item.get("descripcion", "") or "") + " | " + comp_obs,
                        "repuestos": comp_rep, "observaciones": comp_obs
                    })
                    st.success("\u2713 Completado y registrado en hoja de vida")
                    st.rerun()

            # Generar reporte PDF
            if estado == "Completado":
                equipo_data = next((e for e in equipos if e.get("id") == item.get("equipo_id")), {})
                if equipo_data:
                    pdf_bytes = generar_reporte_pdf(equipo_data, item)
                    nombre_eq = item.get("equipo_nombre", "equipo").replace(" ", "_")
                    st.download_button(
                        "\U0001f4c4 Descargar Reporte PDF",
                        pdf_bytes,
                        f"Reporte_{nombre_eq}_{item.get('fecha_ejecucion', date.today())}.pdf",
                        mime="application/pdf",
                        key=f"pdf_{item.get('id')}"
                    )


# ================================================================
# HOJAS DE VIDA
# ================================================================
elif menu == "Hojas de Vida":
    equipos = sb_get(T_EQUIPOS, "order=nombre.asc")
    hojas = sb_get(T_HOJAS, "order=fecha.desc")

    st.markdown(f"<div style='font-family:serif; font-size:20px; color:{COLOR};'>Hojas de Vida de Equipos</div>", unsafe_allow_html=True)

    if equipos:
        opciones_eq = {f"{e.get('nombre')} \u2014 {e.get('marca')} {e.get('modelo')} (Serie: {e.get('serie', 'N/A')})": e.get("id") for e in equipos}
        equipo_sel = st.selectbox("Seleccionar equipo:", list(opciones_eq.keys()), key="hv_eq")
        equipo_id = opciones_eq[equipo_sel]
        equipo_data = next((e for e in equipos if e.get("id") == equipo_id), {})

        st.markdown(f"""
        <div style='background:#111; border:1px solid {COLOR}22; border-radius:6px; padding:14px 18px; margin:8px 0;'>
            <strong style='color:{COLOR};'>{equipo_data.get('nombre')}</strong> \u2014 {equipo_data.get('marca')} {equipo_data.get('modelo')}<br>
            <span style='color:#888; font-size:12px;'>Serie: {equipo_data.get('serie', '\u2014')} | INVIMA: {equipo_data.get('registro_invima', '\u2014')} | Riesgo: {equipo_data.get('clasificacion_riesgo', '\u2014')} | Cliente: {equipo_data.get('cliente_nombre', '\u2014')}</span>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<hr class='divider'>", unsafe_allow_html=True)

        with st.expander("\u2795 Registrar mantenimiento manual"):
            hm1, hm2 = st.columns(2)
            with hm1:
                hv_tipo = st.selectbox("Tipo", ["Preventivo", "Correctivo", "Calibraci\u00f3n", "Inspecci\u00f3n"], key="hv_tipo")
                hv_fecha = st.date_input("Fecha", key="hv_fecha")
                hv_tecnico = st.text_input("T\u00e9cnico", key="hv_tec")
            with hm2:
                hv_trabajo = st.text_area("Trabajo realizado", key="hv_trab")
                hv_repuestos = st.text_input("Repuestos", key="hv_rep")
                hv_obs = st.text_area("Observaciones", key="hv_obs")

            if st.button("\u2713 Registrar", key="btn_hv"):
                sb_post(T_HOJAS, {
                    "equipo_id": equipo_id, "equipo_nombre": equipo_sel.split(" \u2014 ")[0],
                    "tipo_mantenimiento": hv_tipo, "fecha": str(hv_fecha),
                    "tecnico": hv_tecnico, "trabajo_realizado": hv_trabajo,
                    "repuestos": hv_repuestos, "observaciones": hv_obs
                })
                st.success("\u2713 Registrado")
                st.rerun()

        hojas_equipo = [h for h in hojas if h.get("equipo_id") == equipo_id]

        st.markdown(f"<div style='font-family:serif; font-size:14px; color:{COLOR}; margin:12px 0;'>Historial de mantenimiento ({len(hojas_equipo)} registros)</div>", unsafe_allow_html=True)

        if hojas_equipo:
            for h in hojas_equipo:
                tipo_color = {"Preventivo": "#44ff88", "Correctivo": "#ff4444", "Calibraci\u00f3n": "#4488ff", "Inspecci\u00f3n": "#ffaa44"}.get(h.get("tipo_mantenimiento"), "#888")
                st.markdown(f"""
                <div style='background:#0d0d0d; border:1px solid {COLOR}15; border-left:3px solid {tipo_color}; border-radius:4px; padding:12px 16px; margin:6px 0;'>
                    <div style='display:flex; justify-content:space-between;'>
                        <span style='color:#e0e0e0; font-size:13px;'><strong>{h.get('tipo_mantenimiento')}</strong> \u2014 {h.get('fecha', '\u2014')}</span>
                        <span style='color:{tipo_color}; font-size:11px;'>{h.get('tecnico', '\u2014')}</span>
                    </div>
                    <div style='color:#888; font-size:12px; margin-top:4px;'>{h.get('trabajo_realizado', '\u2014')}</div>
                    <div style='color:#666; font-size:11px; margin-top:2px;'>Repuestos: {h.get('repuestos', '\u2014')}</div>
                </div>
                """, unsafe_allow_html=True)

            # Exportar hoja de vida Excel mejorada
            excel_bytes = generar_hoja_vida_excel(equipo_data, hojas_equipo)
            nombre_eq = equipo_data.get("nombre", "equipo").replace(" ", "_")
            st.download_button(
                "\U0001f4cb Exportar Hoja de Vida (Excel)",
                excel_bytes,
                f"Hoja_Vida_{nombre_eq}_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="exp_hv"
            )
        else:
            st.markdown("<div style='text-align:center; color:#555; padding:30px;'>No hay registros para este equipo</div>", unsafe_allow_html=True)
    else:
        st.markdown("<div style='text-align:center; color:#555; padding:40px;'>Registra equipos primero en Inventario</div>", unsafe_allow_html=True)


# ================================================================
# CLIENTES
# ================================================================
elif menu == "Clientes":
    clientes = sb_get(T_CLIENTES, "order=created_at.desc")
    equipos_all = sb_get(T_EQUIPOS, "order=nombre.asc")

    st.markdown(f"<div style='font-family:serif; font-size:20px; color:{COLOR};'>Clientes / Consultorios</div>", unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown(f"<div class='metric-card'><div class='metric-num'>{len(clientes)}</div><div class='metric-label'>CLIENTES</div></div>", unsafe_allow_html=True)
    with col2:
        st.markdown(f"<div class='metric-card'><div class='metric-num'>{len(equipos_all)}</div><div class='metric-label'>EQUIPOS TOTALES</div></div>", unsafe_allow_html=True)

    st.markdown("<hr class='divider'>", unsafe_allow_html=True)

    with st.expander("\u2795 Nuevo cliente"):
        nc1, nc2 = st.columns(2)
        with nc1:
            cl_nombre = st.text_input("Nombre", key="cl_nom")
            cl_contacto = st.text_input("Contacto", key="cl_cont")
            cl_telefono = st.text_input("Tel\u00e9fono", key="cl_tel")
        with nc2:
            cl_email = st.text_input("Email", key="cl_email")
            cl_direccion = st.text_input("Direcci\u00f3n", key="cl_dir")
            cl_nit = st.text_input("NIT", key="cl_nit")

        if st.button("\u2713 Registrar cliente", key="btn_cl"):
            if cl_nombre:
                sb_post(T_CLIENTES, {
                    "nombre": cl_nombre, "contacto": cl_contacto,
                    "telefono": cl_telefono, "email": cl_email,
                    "direccion": cl_direccion, "nit": cl_nit, "estado": "Activo"
                })
                st.success("\u2713 Registrado")
                st.rerun()

    st.markdown("<hr class='divider'>", unsafe_allow_html=True)

    for cliente in clientes:
        eq_count = len([e for e in equipos_all if e.get("cliente_nombre") == cliente.get("nombre")])
        with st.expander(f"{cliente.get('nombre')} \u2014 {eq_count} equipos"):
            st.markdown(f"**Contacto:** {cliente.get('contacto', '\u2014')}")
            st.markdown(f"**Tel\u00e9fono:** {cliente.get('telefono', '\u2014')}")
            st.markdown(f"**Email:** {cliente.get('email', '\u2014')}")
            st.markdown(f"**Direcci\u00f3n:** {cliente.get('direccion', '\u2014')}")
            st.markdown(f"**NIT:** {cliente.get('nit', '\u2014')}")
            num = (cliente.get("telefono") or "").replace(" ", "").replace("-", "")
            if num:
                st.markdown(f"<a href='https://wa.me/57{num}' target='_blank' style='color:{COLOR};'>\U0001f4f1 WhatsApp</a>", unsafe_allow_html=True)


# ================================================================
# ALERTAS
# ================================================================
elif menu == "Alertas":
    alertas = sb_get(T_ALERTAS, "order=fecha_vencimiento.asc")
    hoy = date.today()
    vencidas = [a for a in alertas if a.get("fecha_vencimiento") and a.get("fecha_vencimiento") < str(hoy) and a.get("estado") != "Completada"]

    st.markdown(f"<div style='font-family:serif; font-size:20px; color:{COLOR};'>Alertas y Vencimientos</div>", unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown(f"<div class='metric-card'><div class='metric-num' style='color:#ff4444;'>{len(vencidas)}</div><div class='metric-label'>VENCIDAS</div></div>", unsafe_allow_html=True)
    with col2:
        st.markdown(f"<div class='metric-card'><div class='metric-num'>{len(alertas)}</div><div class='metric-label'>TOTAL</div></div>", unsafe_allow_html=True)

    st.markdown("<hr class='divider'>", unsafe_allow_html=True)

    with st.expander("\u2795 Nueva alerta"):
        al_titulo = st.text_input("T\u00edtulo", key="al_tit")
        al_desc = st.text_area("Descripci\u00f3n", key="al_desc")
        al_fecha = st.date_input("Fecha vencimiento", key="al_fecha")
        al_pri = st.selectbox("Prioridad", ["Normal", "Alta", "Urgente"], key="al_pri")

        if st.button("\u2713 Crear alerta", key="btn_al"):
            if al_titulo:
                sb_post(T_ALERTAS, {
                    "titulo": al_titulo, "descripcion": al_desc,
                    "fecha_vencimiento": str(al_fecha), "prioridad": al_pri, "estado": "Pendiente"
                })
                st.success("\u2713 Creada")
                st.rerun()

    st.markdown("<hr class='divider'>", unsafe_allow_html=True)

    for alerta in alertas:
        if alerta.get("estado") == "Completada":
            continue
        fecha = alerta.get("fecha_vencimiento", "")
        es_vencida = fecha and fecha < str(hoy)

        with st.expander(f"{'\U0001f534' if es_vencida else '\U0001f7e1'} {alerta.get('titulo')} \u2014 {fecha}"):
            st.markdown(f"**Descripci\u00f3n:** {alerta.get('descripcion', '\u2014')}")
            st.markdown(f"**Prioridad:** {alerta.get('prioridad', 'Normal')}")
            if st.button("\u2713 Completada", key=f"al_c_{alerta.get('id')}"):
                sb_patch(T_ALERTAS, alerta.get("id"), {"estado": "Completada"})
                st.rerun()


# ================================================================
# DOCUMENTOS
# ================================================================
elif menu == "Documentos":
    st.markdown(f"<div style='font-family:serif; font-size:20px; color:{COLOR};'>Documentos</div>", unsafe_allow_html=True)

    archivo = st.file_uploader("Subir documento", type=["pdf", "png", "jpg", "jpeg", "doc", "docx", "xlsx"])
    if archivo and st.button("\U0001f4e4 Subir", key="btn_up"):
        url = subir_documento(archivo)
        if url:
            st.success("\u2713 Subido")
            st.markdown(f"[Ver documento]({url})")

    st.markdown("<hr class='divider'>", unsafe_allow_html=True)

    docs = listar_documentos()
    if docs:
        for doc in docs:
            nombre_doc = doc.get("name", "")
            if nombre_doc:
                url_doc = f"{SUPABASE_URL}/storage/v1/object/public/documentos/{SLUG}/{nombre_doc}"
                st.markdown(f"""
                <div style='background:#111; border:1px solid {COLOR}22; padding:10px 16px; border-radius:4px; margin:4px 0; display:flex; justify-content:space-between;'>
                    <span style='color:#e0e0e0; font-size:13px;'>\U0001f4c4 {nombre_doc}</span>
                    <a href='{url_doc}' target='_blank' style='color:{COLOR}; font-size:11px;'>Descargar</a>
                </div>
                """, unsafe_allow_html=True)
    else:
        st.markdown("<div style='text-align:center; color:#555; padding:30px;'>No hay documentos</div>", unsafe_allow_html=True)


# -------- FOOTER --------
st.markdown(f"""
<div style='text-align:center; padding:20px 0; margin-top:30px; border-top:1px solid {COLOR}22;
    font-family:serif; font-size:10px; color:{COLOR}66; letter-spacing:3px;'>
    {NOMBRE_EMPRESA.upper()} \u2014 POWERED BY A.R.I.Z.A.
</div>
""", unsafe_allow_html=True)